"""Tear-sheet export: Excel (.xlsx) and PDF for a single indicative pricing run."""
from __future__ import annotations
from io import BytesIO
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fpdf import FPDF


def _ts() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _safe(text: str) -> str:
    """Replace common non-Latin-1 chars; fpdf2 Helvetica is Latin-1 only."""
    return (text
            .replace("—", "-").replace("–", "-")
            .replace("×", "x").replace("·", ".")
            .replace("…", "...").replace("’", "'")
            .encode("latin-1", errors="replace").decode("latin-1"))


def _fmt(v) -> str:
    if isinstance(v, float):
        return f"{v:,.4f}" if abs(v) < 10_000 else f"{v:,.2f}"
    if isinstance(v, int):
        return f"{v:,}"
    return str(v)


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────
_NAVY  = "1A2744"
_BDARK = "0F3460"
_ACCNT = "4FC3F7"
_DKBG  = "111827"


def build_excel(product_name: str, inputs: dict, result: dict, xva: dict) -> bytes:
    """Return xlsx bytes ready for st.download_button."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tear Sheet"

    f_ttl = Font(bold=True, color=_ACCNT, size=13)
    f_prd = Font(bold=True, color="FFFFFF", size=11)
    f_ts  = Font(color="AAAAAA", size=8, italic=True)
    f_sec = Font(bold=True, color=_ACCNT, size=9)
    f_lbl = Font(color="BBBBBB", size=9)
    f_val = Font(bold=True, color="FFFFFF", size=9)
    f_dis = Font(color="777777", size=7, italic=True)

    fl_navy = PatternFill("solid", fgColor=_NAVY)
    fl_sec  = PatternFill("solid", fgColor=_BDARK)
    fl_row  = PatternFill("solid", fgColor=_DKBG)
    al      = Alignment(horizontal="left", vertical="center", indent=1)

    def put(row, col, val, font=None, fill=None, merge_to=None):
        if merge_to:
            from openpyxl.utils import get_column_letter
            ws.merge_cells(
                f"{get_column_letter(col)}{row}:{get_column_letter(merge_to)}{row}"
            )
        c = ws.cell(row=row, column=col, value=val)
        if font: c.font = font
        if fill: c.fill = fill
        c.alignment = al
        return c

    r = 1
    ws.row_dimensions[1].height = 26
    put(r, 1, "Macro QIS Workbench — Indicative Pricing", f_ttl, fl_navy, 4); r += 1
    ws.row_dimensions[2].height = 20
    put(r, 1, product_name, f_prd, fl_navy, 4); r += 1
    put(r, 1, _ts(), f_ts, fl_navy, 4); r += 2

    def section(title: str, data: dict):
        nonlocal r
        ws.row_dimensions[r].height = 18
        put(r, 1, f"  {title}", f_sec, fl_sec, 4); r += 1
        for label, value in data.items():
            ws.row_dimensions[r].height = 15
            put(r, 1, str(label), f_lbl, fl_row)
            put(r, 2, _fmt(value) if isinstance(value, (int, float)) else str(value),
                f_val, fl_row)
            r += 1
        r += 1

    section("INPUTS", inputs)
    _skip = {"product", "legs"}
    section("PRICING RESULTS",
            {k: v for k, v in result.items()
             if k not in _skip and not isinstance(v, (dict, list))})
    section("xVA SUMMARY",
            {k: v for k, v in xva.items()
             if not isinstance(v, (dict, list))})
    put(r + 1, 1,
        "DISCLAIMER: Indicative only. Not a commitment to trade. "
        "PoC model — not for production use.",
        f_dis, merge_to=4)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────────────────────────────────────

class _PDF(FPDF):
    def footer(self):
        self.set_y(-20)
        self.set_fill_color(245, 245, 245)
        self.rect(10, self.get_y(), 190, 14, "F")
        self.set_xy(12, self.get_y() + 1.5)
        self.set_font("Helvetica", "I", 6.5)
        self.set_text_color(120, 120, 120)
        self.multi_cell(
            186, 3.5,
            "DISCLAIMER: This document is indicative only and does not constitute a commitment "
            "to trade. Produced by Macro QIS Workbench (Proof of Concept) - not validated "
            "for production or client distribution.",
        )


def build_pdf(product_name: str, inputs: dict, result: dict, xva: dict) -> bytes:
    """Return PDF bytes ready for st.download_button."""
    pdf = _PDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=25)

    # masthead
    pdf.set_fill_color(26, 39, 68)
    pdf.rect(0, 0, 210, 40, "F")
    pdf.set_xy(10, 8)
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(79, 195, 247)
    pdf.cell(190, 8, "Macro QIS Workbench - Indicative Pricing")
    pdf.set_xy(10, 19)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(190, 7, _safe(product_name))
    pdf.set_xy(10, 29)
    pdf.set_font("Helvetica", "I", 7.5)
    pdf.set_text_color(170, 170, 170)
    pdf.cell(190, 5, f"{_ts()}   |   PoC - Indicative Only")

    def section_at(title: str, data: dict, x: float, w: float, y: float) -> float:
        """Write a section at absolute (x, y); return the final y."""
        lw = w * 0.57
        vw = w * 0.43
        pdf.set_xy(x, y)
        pdf.set_fill_color(15, 52, 96)
        pdf.set_text_color(79, 195, 247)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(w, 6.5, _safe(f"  {title}"), fill=True)
        y += 6.5
        for label, value in data.items():
            pdf.set_xy(x, y)
            pdf.set_font("Helvetica", "", 7.5)
            pdf.set_text_color(90, 90, 90)
            pdf.cell(lw, 5.5, _safe(f"  {label}"))
            pdf.set_xy(x + lw, y)
            pdf.set_font("Helvetica", "B", 7.5)
            pdf.set_text_color(30, 30, 30)
            vs = (_fmt(value) if isinstance(value, (int, float)) else _safe(str(value)))
            pdf.cell(vw, 5.5, vs)
            y += 5.5
        return y + 3

    _skip = {"product", "legs"}
    res_disp = {k: v for k, v in result.items()
                if k not in _skip and not isinstance(v, (dict, list))}
    xva_disp = {k: v for k, v in xva.items()
                if not isinstance(v, (dict, list))}

    Y0 = 47.0
    y_l = section_at("INPUTS",      inputs,   x=10,  w=88, y=Y0)
    y_l = section_at("xVA SUMMARY", xva_disp, x=10,  w=88, y=y_l)
    y_r = section_at("PRICING RESULTS", res_disp, x=112, w=88, y=Y0)
    pdf.set_y(max(y_l, y_r) + 2)

    buf = BytesIO()
    pdf.output(buf)
    return buf.getvalue()
